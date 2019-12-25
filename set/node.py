from message import fail, initContractPropose, contractPropose, contractSelect, LockedTransfer, sendSecret, resSecret, \
    revealSecret, resRevealSecret, completePayGo, waitContractSelect, selectionComplete, contractConfirm, contractReject, endcontractPropose
from algorism import contract_bundle, RTT
from settingParameter import on_chain_access_cost
from structure import contractTable, payer_result
from raiden.utils import sha3
from settingParameter import alpha, contract_meaningful_incentive_constant, contract_meaningful_delay_constant, \
    time_meaningful_constant, theta
from contract import register_secret, close_channel, update_NonClosingBalanceProof, settle_channel, unlock
from raiden.utils.signing import pack_data

from openpyxl import Workbook
import os, math, time
# import threading


class Node:
    def __init__(self, w3, account, name, line, result,lock):
        self.w3 = w3
        self.name = name
        self.line = line
        self.account = account
        self.address = account.address
        self.private_key = account.privateKey
        self.cr = 0
        self.channel_state = []
        self.partner = {}
        self.contract_table = {}
        self.experiment_result = result
        self.probability = {}
        self.RTT = RTT()
        self.lock = lock

    def get_channel_probability(self):
        channel_probability = {}
        self.lock.acquire()
        try:
            probability = self.probability
        finally:
            self.lock.release()
        if 0 == len(probability.keys()):
            for i in range(theta):
                channel_probability[i] = round(100 / theta, 4)
            self.probability = channel_probability
        else:
            channel_probability = probability

        # print("channel_probability : ", channel_probability)

        return channel_probability

    def create_channel_state(self, partner, channel_state):
        self.channel_state.append(channel_state)
        # self.partner[channel_state.addrs[1-channel_state.i]] = channel_state
        self.partner[partner] = channel_state

    def init_contract_propose(self, round, initiator, target, amount, secret, start_time, Omega, Omega_prime, payment_state, start):
        cr = '0x' + sha3(secret).hex()
        channel_probability = self.get_channel_probability()
        bundle = contract_bundle(initiator.address).execute(channel_probability, Omega, Omega_prime)
        M = []
        success = 0
        for recipient in initiator.partner.keys() :
            if self.partner[recipient].check_average_capacity(amount):
                success +=1
                message = contractPropose(cr, initiator, self, recipient, target, amount, bundle)

                if not cr in self.contract_table.keys():
                    self.experiment_result.payGo_startTime = start
                    self.contract_table[cr] = contractTable(message)
                    self.contract_table[cr].secret = secret
                    self.contract_table[cr].payment_startTime = start_time
                    self.experiment_result.protocol_time[round] = {"propose": start_time, "select": 0, "lockTransfer": 0}
                    self.experiment_result.select_time[round] = {}

                # print("{} -> {} init_contract_propose".format(initiator.name, recipient.name))\
                self.contract_table[cr].contract_bundle[recipient] = bundle

                M.append(message)

        if success == 0 :
            print("initiaotr error", self.name)
            M = self.init_contract_propose(round, initiator, target, amount, secret, start_time, payment_state, start)

        return M

    def send_contract_propose(self, cr, initiator, partner, target, amount, bundle):
        self.contract_table[cr].check_propose.append(partner)
        message = contractPropose(cr, initiator, self, partner, target, amount, bundle)
        self.contract_table[message.cr].contract_bundle[partner] = bundle

        return message

    def receive_contract_propose(self, message, payment_state, count, Omega, Omega_prime, round):
        # print("[{}] propose {} -> {}".format(round, message.producer.name, self.name))
        M = []
        if not message.cr in self.contract_table.keys():
            self.contract_table[message.cr] = contractTable(message)
        self.contract_table[message.cr].receive_contract_bundle[message.producer] = message.contract_bundle

        if count > self.contract_table[message.cr].selection_direction:
            self.contract_table[message.cr].selection_direction += 1
            return M

        self.contract_table[message.cr].selection_direction = 1

        if payment_state == "contractPropose":
            if self.address == message.target.address:
                for recipient in self.contract_table[message.cr].receive_contract_bundle.keys():
                    if not recipient in self.contract_table[message.cr].check_propose :
                        M.append(self.send_target_contract_select(message.cr, recipient, 0, 0, 0, 0))
            else:
                success = 0
                create_bundle = {}
                for partner in self.partner.keys():
                    possibility = self.contract_select(message.cr, partner, message.amount, message.producer,
                                                       self.contract_table[message.cr].receive_contract_bundle, False, "propose")
                    if partner.line != message.producer.line and possibility != -1 and\
                       not partner in self.contract_table[message.cr].check_propose :
                        success +=1
                        channel_probability = self.get_channel_probability()
                        if partner == message.target:
                            create_bundle = {"Incentive": 1, "Delay": 5}
                        elif len(create_bundle) ==0:
                            create_bundle = contract_bundle(self.address).execute(channel_probability, Omega, Omega_prime)
                        M.append(self.send_contract_propose(message.cr, message.initiator, partner, message.target,
                                                            message.amount, create_bundle))
                if success == 0 :
                    # print("{} propose X {}".format(message.initiator.name, self.name))
                    M.append(message)
        return M

    def send_target_contract_select(self, cr, recipient, incentive, delay, additional_incentive, additional_delay):
        table = self.contract_table[cr]
        self.contract_table[cr].check_propose.append(recipient)

        selected_incentive = table.receive_contract_bundle[recipient]["Incentive"]
        selected_delay = table.receive_contract_bundle[recipient]["Delay"]

        additional_incentive += incentive
        additional_delay += delay
        new_time = time.time()
        message = contractSelect(cr, table.initiator, self, recipient, table.target,
                                 selected_incentive, selected_delay, 0, additional_incentive,
                                 additional_delay, new_time)
        # path 설정
        message.path.append(self.name)

        return message

    def send_contract_select(self, cr, producer, consumer, incentive, delay, additional_incentive, additional_delay,
                             selected_index, path, propose_endTime):
        # 14 select
        table = self.contract_table[cr]
        self.contract_table[cr].state = "contractSelect"
        self.contract_table[cr].send_selected_contract.append(producer)
        try:
            selected_incentive = int(round(table.receive_contract_bundle[producer]["Incentive"][selected_index][0]))
            selected_delay = int(round(table.receive_contract_bundle[producer]["Delay"][selected_index]))
        except:
            print("contract error, receive contract bundle : {}".format(table.receive_contract_bundle[producer]))
            print("contract error selected index : ", selected_index)
            length = len(table.receive_contract_bundle[producer]["Incentive"])
            selected_incentive = int(round(table.receive_contract_bundle[producer]["Incentive"][length - 1][0]))
            selected_delay = int(round(table.receive_contract_bundle[producer]["Delay"][length - 1]))

        if selected_incentive >= on_chain_access_cost:
            print("incentive 초과 : {}, selected_index : {}, bundle : {}".format(selected_incentive, selected_index, table.receive_contract_bundle[producer]["Incentive"]))
            message = fail("onchainCostExcess")
        else:
            additional_incentive += incentive
            additional_delay += delay
            message = contractSelect(cr, table.initiator, self, producer, table.target, selected_incentive,
                                     selected_delay,
                                     selected_index, additional_incentive, additional_delay, propose_endTime)

            # path 설정
            message.path.append(self.name)

        return message

    def contract_select(self, cr, selected_consumer, amount, producer, contract_bundle, already_section, type):
        index = -1
        # queueing_record [0] : amount, [1] : delay
        queueing_record = self.partner[selected_consumer].get_queueing_record()

        if len(queueing_record) == 0:
            length = len(contract_bundle[producer]["Delay"])
            self.contract_table[cr].aready_reservation_delay \
                = contract_bundle[producer]["Delay"][length-1] / contract_meaningful_delay_constant
            return len(contract_bundle[producer]["Delay"]) - 1

        if already_section :
            delay = self.contract_table[cr].aready_reservation_delay
        else :
            capacity = self.partner[selected_consumer].get_average_capacity()

            if capacity == 0: return index

            # margin = (queueing_record[1] * (amount / capacity) / queueing_record[0]) / 3
            margin = 0

            delay = queueing_record[1] * (amount / capacity) / queueing_record[0] + margin
            # print("queueing_record[1]", queueing_record[1])
            # print("(amount / capacity)", (amount / capacity))
            # print("queueing_record[0]", queueing_record[0])
            # print("margin : ", margin)
            # print("deay :", delay)
            # print("exclude margin delay :", delay-margin)
            # print("contract :", (contract_bundle[producer]["Delay"]))
        for i in range(len(contract_bundle[producer]["Delay"]) - 1, -1, -1):
            # delay * 5 -> 수정 가능
            if delay <= contract_bundle[producer]["Delay"][i] / contract_meaningful_delay_constant:
                index = i
                self.contract_table[cr].contract_predict_time = \
                    contract_bundle[producer]["Delay"][i] / contract_meaningful_delay_constant

                if type == "selection" and not already_section :
                    self.contract_table[cr].aready_reservation_delay = delay
                    self.partner[selected_consumer].update_average_capacity_awaitAmount(cr,amount, time.time())
                break
        return index

    def receive_contract_select(self, message, count, payment_state, Omega, r):
        # print("[{}, {}]{} <- {} send_contract_select".format(message.initiator.name, r, self.name,
        #                                                      message.consumer.name))

        M = []
        if self.contract_table[message.cr].state == "contractSelect":
            return M

        if payment_state == "lockedTransfer":
            return M

        if not message.consumer in self.contract_table[message.cr].temp_consumer:
            self.contract_table[message.cr].temp_selected_contract[message.consumer] = message.incentive, message.delay
            self.contract_table[message.cr].temp_additional_contract[
                message.consumer] = message.additional_incentive, message.additional_delay
            self.contract_table[message.cr].temp_selected_index[message.consumer] = message.selected_index
            self.contract_table[message.cr].temp_consumer.append(message.consumer)
        # else :
        #     print("[{}, {}]{}' with {} balance : {}, reserve amount : {}".format(
        #         message.initiator.name, r, self.name, message.consumer.name, self.partner[message.consumer].test_get_balance(),
        #         self.partner[message.consumer].get_reserve_amount()
        #     ))

        if count > self.contract_table[message.cr].selection_direction:
            self.contract_table[message.cr].selection_direction += 1
            return M

        # 나중을 위해 초기화
        self.contract_table[message.cr].selection_direction = 1

        table = self.contract_table[message.cr]

        if message.initiator != self :
            temp_consumer = []
            recipient = list(self.contract_table[message.cr].receive_contract_bundle.keys())[0]
            for i in range(len(table.temp_consumer)):
                selected_incentive = table.temp_selected_contract[table.temp_consumer[i]][0]
                additional_incentive = table.temp_additional_contract[table.temp_consumer[i]][0]
                amount = selected_incentive + additional_incentive + table.amount

                selected_index = self.contract_select(message.cr, table.temp_consumer[i], amount,
                                                      recipient, table.receive_contract_bundle, False, "check")

                if selected_index != -1 : temp_consumer.append(table.temp_consumer[i])

            # if len(table.temp_consumer) != len(temp_consumer) :
            #     print("len : {} : {}".format(len(table.temp_consumer), len(temp_consumer)))

            table.temp_consumer = temp_consumer

            if len(temp_consumer) == 0:
                M.append(message)
                # print("[{}] {} initiator channel bankrupt".format(r, self.name))
                return M

        # if target
        selected_consumer = self.contract_choice_Maxium_U(table, Omega)

        selected_contract = table.temp_selected_contract[selected_consumer]
        additional_contract = table.temp_additional_contract[selected_consumer]
        selected_index = table.temp_selected_index[selected_consumer]

        self.contract_table[message.cr].selected_contract = selected_contract
        self.contract_table[message.cr].additional_contract = additional_contract
        self.contract_table[message.cr].selected_index = selected_index
        self.contract_table[message.cr].consumer = selected_consumer

        if message.consumer != message.target:
            probability = self.get_channel_probability()
            if probability[0] >= 1.8 :
                for i in table.temp_selected_index.values() :
                    index = theta - len(self.contract_table[message.cr].contract_bundle[message.consumer]["Delay"]) + i
                    probability[index] += 0.1

                    value = round((0.1 / (theta -1)), 5)
                    rest = 0.1 - (value * (theta -1))
                    for j in range(len(probability)) :
                        if j != index :
                            probability[j] = round(probability[j] - value, 5)
                    probability[index] = round(probability[index]- rest, 5)

                self.lock.acquire()
                try:
                    self.probability = probability
                finally:
                    self.lock.release()

        amount = self.contract_table[message.cr].amount + selected_contract[0] + additional_contract[0]

        if self == message.initiator:
            # a = self.reserve_payment(message.cr, selected_consumer, amount)
            self.partner[selected_consumer].update_average_capacity(-amount)
            new_time = time.time()
            self.contract_table[message.cr].state = "contractSelect"
            temp = self.experiment_result.protocol_time[r]["propose"]
            self.experiment_result.protocol_time[r]["propose"] = message.propose_endTime - temp
            self.experiment_result.protocol_time[r]["select"] =  new_time- message.propose_endTime
            M.append(self.send_secret(message.cr, message.initiator, message.target))
            M.append(self.send_contract_confirm(message.cr, selected_consumer, selected_contract[0] ,selected_contract[1]))
        else:
            sucess = 0
            aready_reservation = False
            capacity = self.partner[selected_consumer].get_average_capacity()
            count = self.partner[selected_consumer].update_awaitAmount(self.RTT.get_accumulate_time())
            if capacity == 0 :
                M.append(message)
                return M
            self.contract_table[message.cr].amount_ratio = amount / capacity
            for recipient in self.contract_table[message.cr].receive_contract_bundle.keys():
                selected_index = self.contract_select(message.cr, selected_consumer, amount,
                                                          recipient, table.receive_contract_bundle, aready_reservation, "selection")

                if selected_index == -1:
                    continue
                else:
                    new_time = time.time()
                    if not aready_reservation :
                        aready_reservation = True
                        new_capacity = self.partner[selected_consumer].get_average_capacity()
                        # print("test ", self.partner[selected_consumer].get_wait_confirm(self.RTT.get_accumulate_time()))
                        self.contract_table[message.cr].selection_RTT_start = new_time
                        temp1 = {self.name: new_capacity}
                        temp2 = {self.name: self.partner[selected_consumer].test_get_balance()}
                        temp3 = {self.name: table.receive_contract_bundle[recipient]["Delay"][selected_index] / contract_meaningful_delay_constant}

                        if not r in message.initiator.experiment_result.capacity :
                            message.initiator.experiment_result.capacity[r] = {}
                            message.initiator.experiment_result.balance[r] = {}
                            message.initiator.experiment_result.selected_contract[r] = {}

                        message.initiator.experiment_result.capacity[r].update(temp1)
                        message.initiator.experiment_result.balance[r].update(temp2)
                        message.initiator.experiment_result.selected_contract[r].update(temp3)

                        temp = {self.name: self.partner[selected_consumer].get_wait_confirm_count()}
                        if not r in message.initiator.experiment_result.pending_queue:
                            message.initiator.experiment_result.pending_queue[r] = {}
                        message.initiator.experiment_result.pending_queue[r].update(temp)
                        # if self.partner[selected_consumer].get_wait_confirm_count() >= 2:
                            # print("pending : ", self.partner[selected_consumer].get_wait_confirm_count())

                        # print("[{}, {}] probility : {}".format(message.initiator.name, self.name, self.get_channel_probability()))
                        # print("[{}, {}] contract : {}".format(message.initiator.name, self.name, table.receive_contract_bundle[recipient]))

                    sucess +=1
                    M.append(self.send_contract_select(message.cr, recipient, message.consumer, selected_contract[0],
                                                       selected_contract[1], additional_contract[0],
                                                       additional_contract[1], selected_index, message.path,
                                                       message.propose_endTime))
                    # print("[{}. {}] select {}".format(message.initiator, r, selected_index))
            if sucess == 0:
                M.append(message)

        return M

    def contract_choice_short_pass(self, table):
        select = 10000000000  # infinite value
        selected_consumer = ""
        for i in range(len(table.temp_consumer)):
            selected_delay = table.temp_selected_contract[table.temp_consumer[i]][1]
            additional_delay = table.temp_additional_contract[table.temp_consumer[i]][1]
            if selected_delay + additional_delay < select:
                select = selected_delay + additional_delay
                selected_consumer = table.temp_consumer[i]

        return selected_consumer


    def contract_choice_Maxium_U(self, table, Omega):
        utility = -1000000  # infinite value
        selected_consumer = ""

        for i in range(len(table.temp_consumer)):
            selected_contract = table.temp_selected_contract[table.temp_consumer[i]]
            temp = contract_bundle(self.address).get_producer_utility((selected_contract[0]) / contract_meaningful_incentive_constant,
                                                                    (selected_contract[1]) / contract_meaningful_delay_constant, Omega)

            # print("{} incentive : ({},{}), delay : ({},{}) utility : {}".format(
            #     i,selected_contract[0] / contract_meaningful_incentive_constant , additional_contract[0] / contract_meaningful_incentive_constant,
            #     selected_contract[1] / contract_meaningful_delay_constant , additional_contract[1] / contract_meaningful_delay_constant, temp))
            if temp > utility:
                utility = temp
                selected_consumer = table.temp_consumer[i]
        # print("choice utility :", utility)
        # print("choice node :", selected_consumer.name)
        return selected_consumer


    def send_contract_confirm(self, cr, selected_consumer, selected_incentive, selected_delay):
        message = contractConfirm(cr, self, selected_consumer, selected_incentive, selected_delay)

        return message

    def send_contract_reject(self, cr, recipent, id):
        message = contractReject(cr, self, recipent, id)

        return message

    def receive_contract_reject(self, message, round):
        # print("receive_contract_reject", self.name, message.type)
        M = []
        table = self.contract_table[message.cr]
        consumer = table.consumer
        result_incentive = (table.selected_contract[0] + table.additional_contract[0])
        if message.type == "contractRejectDown" :
            self.partner[consumer].update_average_capacity(table.amount + result_incentive)
            if consumer != table.target :
                M.append(self.send_contract_reject(message.cr, consumer, "contractRejectDown"))

        elif message.type == "contractRejectUp" :
            self.partner[consumer].fail_payment(message.cr, table.amount + result_incentive, 0)
            self.partner[consumer].update_average_capacity(table.amount + result_incentive)
            if self != table.initiator :
                pending_payment = self.partner[self.contract_table[message.cr].producer].get_pending_payment()
                try :
                    self.partner[self.contract_table[message.cr].producer].fail_payment(message.cr, pending_payment[message.cr][0], 1)
                except :
                    print("tset : contractRejectUp",pending_payment)
                self.partner[self.contract_table[message.cr].producer].pop_pending_payment(message.cr)
                M.append(self.send_contract_reject(message.cr, self.contract_table[message.cr].producer, "contractRejectUp"))

        return M

    def receive_contract_confirm(self, message, round):
        # print("[{}, {}] {}' receive_contract_confirm ".format(round, self.contract_table[message.cr].initiator.name, self.name))
        M = []
        consumer = self.contract_table[message.cr].consumer
        incentive = self.contract_table[message.cr].selected_contract[0] + \
                    self.contract_table[message.cr].additional_contract[0]
        amount = self.contract_table[message.cr].amount + incentive
        # a = self.reserve_payment(message.cr, consumer, amount)

        wait_amount = self.partner[consumer].get_wait_confirm()
        if not message.cr in wait_amount :
            self.partner[consumer].update_average_capacity(-amount)

        self.contract_table[message.cr].my_contract = message.incentive, message.delay

        if consumer != self.contract_table[message.cr].target :
            M.append(self.send_contract_confirm(message.cr, consumer,
                                                self.contract_table[message.cr].selected_contract[0],
                                                self.contract_table[message.cr].selected_contract[1]))

        if message.cr in self.partner[consumer].wait_confirm[self.partner[consumer].i] :
            self.partner[consumer].wait_confirm[self.partner[consumer].i].pop(message.cr)

        self.RTT.update_PTT(self.name, time.time() - self.contract_table[message.cr].selection_RTT_start)

        return M


    def send_secret(self, cr, initiator, target):
        secret = self.contract_table[cr].secret
        # send Secret
        message = sendSecret(cr, secret, initiator, target)
        return message

    # message.cr, initiator, 0, recipient, target, amount, sha3(message.secret), 0
    def send_locked_transfer(self, cr, initiator, producer, recipient, target, amount, secretHash, receive_BP,
                             receive_selected_contract, receive_aux_contract, r):
        table = self.contract_table[cr]
        result_incentive = (table.selected_contract[0] + table.additional_contract[0])
        result_delay = table.selected_contract[1] + table.additional_contract[1]
        selected_index = table.selected_index
        aux_incentive = []
        aux_delay = []

        if selected_index <= 0:
            for i in range(11) :
                aux_incentive.append(result_incentive)
                aux_delay.append(result_delay)
        else:
            index = selected_index
            for i in range(11):
                if index == 0:
                    aux_incentive.append(int(round(table.contract_bundle[recipient]["Incentive"][index][0]) +
                                        table.additional_contract[0]))
                    aux_delay.append(int(round(table.contract_bundle[recipient]["Delay"][index]) +
                                        table.additional_contract[1]))
                    index +=1
                else:
                    aux_incentive.append(int(round(table.contract_bundle[recipient]["Incentive"][index - 1][0]) +
                                        table.additional_contract[0]))
                    aux_delay.append(int(round(table.contract_bundle[recipient]["Delay"][index - 1]) +
                                         table.additional_contract[1]))
                index -= 1
        # print("{} : selected : {}".format(self.name, receive_selected_contract))
        # print("{} : aux : {}".format(self.name, receive_aux_contract))

        new_time = time.time()
        if self.partner[recipient].check_balance(self.contract_table[cr].amount + result_incentive):
            self.contract_table[cr].lock_time.append(new_time)

            startTime = int(new_time * time_meaningful_constant)
            self.contract_table[cr].startTime = startTime

            # balance data 갱신
            expiration = startTime + result_delay + alpha
            BP = self.partner[recipient].create_BP(self.w3, cr, initiator.address, target.address, secretHash,
                                                       amount, expiration,
                                                       (result_incentive, result_delay), (aux_incentive, aux_delay),
                                                       startTime)
            self.contract_table[cr].aux_incentive = aux_incentive
            self.contract_table[cr].aux_delay = aux_delay

            # print("[{}] bp : {}".format(self.name, self.partner[recipient].test()))

            message = LockedTransfer(cr, self, recipient, BP, target, initiator,
                                     (result_incentive, result_delay),(aux_incentive, aux_delay))


            if r in initiator.experiment_result.pending_payment_settle:
                if self.line in initiator.experiment_result.pending_payment_settle[r] :
                    temp = initiator.experiment_result.pending_payment_settle[r][self.line]["settle_time"]
                    initiator.experiment_result.pending_payment_settle[r][self.line]["settle_time"] = new_time - temp
        else:
            if not r in initiator.experiment_result.pending_payment_settle :
                initiator.experiment_result.onchain_access_pendingTime[r] = time.time()
                initiator.experiment_result.pending_payment_settle[r] = {}
                initiator.experiment_result.pending_payment_settle[r][self.line] = \
                    {"count": len(recipient.partner[self].get_pending_payment()), "settle_time": new_time,
                     "test1" : len(self.partner[recipient].get_pending_payment()), "test2" : self.partner[recipient].get_wait_confirm_count(),
                     "test3" : recipient.partner[self].get_wait_confirm_count()}
            elif not self.line in initiator.experiment_result.pending_payment_settle[r] :
                initiator.experiment_result.onchain_access_pendingTime[r] = time.time()
                initiator.experiment_result.pending_payment_settle[r][self.line] = \
                    {"count": len(recipient.partner[self].get_pending_payment()), "settle_time": new_time,
                     "test1": len(self.partner[recipient].get_pending_payment()),
                     "test2": self.partner[recipient].get_wait_confirm_count(),
                     "test3": recipient.partner[self].get_wait_confirm_count()}

            if self == initiator:
                message = resSecret(cr, initiator, table.secret)
            else:
                pending_payment = self.partner[producer].get_pending_payment()
                # if (receive_aux_contract[1][10] / contract_meaningful_delay_constant) * 5< \
                #         (int(new_time * time_meaningful_constant) - pending_payment[cr][2]) / time_meaningful_constant + 0.008  :
                if 20 < (int(new_time * time_meaningful_constant) - pending_payment[cr][2]) / time_meaningful_constant + 0.008  :
                    self.partner[producer].pop_pending_payment(cr)
                    self.partner[producer].fail_payment(cr, self.contract_table[cr].amount + result_incentive, 1)
                    self.partner[recipient].update_average_capacity(self.contract_table[cr].amount + result_incentive)
                    message = "cancel"
                    initiator.experiment_result.onchain_access +=1

                    initiator.experiment_result.onchain_access_node[r] = self.name
                    initiator.experiment_result.onchain_access_capacity[r] = self.partner[recipient].get_average_capacity()
                    initiator.experiment_result.onchain_access_balance[r] = self.partner[recipient].test_get_balance()
                    initiator.experiment_result.onchain_access_selected_contract[r] = \
                        initiator.experiment_result.selected_contract[r][self.name], receive_aux_contract[1][10] / contract_meaningful_delay_constant
                    temp = initiator.experiment_result.onchain_access_pendingTime[r]
                    initiator.experiment_result.onchain_access_pendingTime[r] = time.time() - temp
                    initiator.experiment_result.onchain_access_pendingQueue[r] = initiator.experiment_result.pending_queue[r][self.name]

                    # delay = (aux_delay[7] / contract_meaningful_delay_constant)
                    # self.update_contract_selection_standard(cr, delay, recipient, 0)
                else :
                    message = LockedTransfer(cr, producer, self, receive_BP, target, initiator,
                                             receive_selected_contract,receive_aux_contract)
        return message

    def cancel_payment(self, cr):
        # [0]:locked_amount [1]:selected_delay [2]:startTime
        # not initiator
        M = []
        if self.contract_table[cr].consumer != self.contract_table[cr].target:
            M.append(self.send_contract_reject(cr, self.contract_table[cr].consumer, "contractRejectDown"))

        M.append(self.send_contract_reject(cr, self.contract_table[cr].producer, "contractRejectUp"))

        return M


    def receive_locked_transfer(self, message, round):
        if not message.cr in self.partner[message.producer].get_pending_payment().keys():
            pending_payment = self.partner[message.producer].locked_BP(message.BP)
            self.partner[message.producer].add_pending_payment(message.cr, pending_payment)
            self.contract_table[message.cr].producer = message.producer
            self.contract_table[message.cr].lock_time.append(time.time())

            if self != message.target :
                temp = {self.line : message.initiator.experiment_result.capacity[round][self.name]}
                temp2 = {self.line: message.initiator.experiment_result.balance[round][self.name]}
                temp3 = {self.line: message.initiator.experiment_result.selected_contract[round][self.name]}
                temp4 = {self.line: message.initiator.experiment_result.pending_queue[round][self.name]}

                message.initiator.experiment_result.capacity[round].update(temp)
                message.initiator.experiment_result.balance[round].update(temp2)
                message.initiator.experiment_result.selected_contract[round].update(temp3)
                message.initiator.experiment_result.pending_queue[round].update(temp4)

        if self.address == message.target.address:
            M = [self.target_reveal_secret(message.cr, message.producer, message.consumer)]
            return M
        else:
            partner = self.contract_table[message.cr].consumer
            amount = self.contract_table[message.cr].amount
            secrethash = message.BP.message_data.secrethash
            M = []
            m = self.send_locked_transfer(message.cr, message.initiator, message.producer, partner, message.target, amount,
                                          secrethash, message.BP, message.selected_contract, message.aux_contract, round)
            if m != "cancel" :
                M.append(m)
            else :
                M = self.cancel_payment(message.cr)

            return M

    def receive_secret(self, message):
        self.contract_table[message.cr].secret = message.secret

        # print("{} -> {} send secret".format(message.initiator.name, self.name))

        M = [self.send_res_secret(message.cr, message.initiator, message.secret)]
        return M

    def send_res_secret(self, cr, initiator, secret):
        message = resSecret(cr, initiator, secret)
        return message

    def recevie_res_secret(self, message, round):
        initiator = message.initiator
        recipient = self.contract_table[message.cr].consumer
        target = self.contract_table[message.cr].target
        amount = self.contract_table[message.cr].amount
        # print("round : ", round)
        new_time = time.time()
        self.experiment_result.protocol_time[round]["lockTransfer"] = new_time
        M = [self.send_locked_transfer(message.cr, initiator, 0, recipient, target, amount, sha3(message.secret),0,
                                       (0,0), (0,0), round)]
        return M

    def target_reveal_secret(self, cr, producer, consumer):
        if self.contract_table[cr].secret != 0:
            new_time = time.time()
            endTime = int(new_time * time_meaningful_constant)
            # print("test endtime :", endTime)
            packed_endTime = pack_data(['uint256'], [endTime])
            hash_endTime = '0x' + sha3(packed_endTime).hex()
            signed_endTime = self.w3.eth.account.signHash(message_hash=hash_endTime, private_key=self.private_key)

            message = revealSecret(cr, self.contract_table[cr].secret, endTime, signed_endTime['signature'].hex(),
                                   producer, consumer, new_time)

            self.contract_table[cr].endTime = endTime
            self.contract_table[cr].signed_endTime = signed_endTime['signature'].hex()

            return message

    def reveal_secret(self, cr, producer, consumer):
        if self.contract_table[cr].secret != 0 or self.contract_table[cr].signed_endTime != 0:
            endTime = self.contract_table[cr].endTime
            signed_endTime = self.contract_table[cr].signed_endTime
            message = revealSecret(cr, self.contract_table[cr].secret, endTime, signed_endTime, producer, consumer)

            print("endTime : ", endTime)
            print("signed_endTime : ", signed_endTime)

            return message

    def receive_reveal_secret(self, message, round, Omega, Omega_prime, total_round):
        self.contract_table[message.cr].secret = message.secret
        self.contract_table[message.cr].endTime = message.endTime
        self.contract_table[message.cr].signed_endTime = message.signed_endTime
        aux_incentive = self.contract_table[message.cr].aux_incentive
        aux_delay = self.contract_table[message.cr].aux_delay

        initiator = self.contract_table[message.cr].initiator
        BP, finalAmount, final_incentive, final_delay = self.partner[message.consumer].unlock(self.w3, message.cr,
                                                                                              message.secret,
                                                                                              message.endTime,
                                                                                              aux_incentive, aux_delay, self,initiator)
        final_incentive = final_incentive / contract_meaningful_incentive_constant
        # print("final_incentive", final_incentive)
        # print("final_delay", final_delay)
        M = [resRevealSecret(message.cr, message.producer, message.consumer, BP, finalAmount)]
        producer = self.contract_table[message.cr].producer
        # print("test ",message.producer.name)
        # print("test2 ",message.consumer.name)

        # print("{} -> {} reveal_secret".format(message.consumer.name, self.name))
        # print("[{}, {}]{} end time : {}, start time {}".format(initiator.name, round, self.name,
        #                                                       message.endTime, self.contract_table[message.cr].startTime))

        if initiator != self:
            M.append(revealSecret(message.cr, self.contract_table[message.cr].secret, message.endTime,
                                  message.signed_endTime, producer, self, message.locked_transfer_endtime))
            if self.line - initiator.line == 1 or self.line - initiator.line == -1:
                self.experiment_result.h1_contract_bundle.append(self.contract_table[message.cr].receive_contract_bundle[producer])
                self.experiment_result.h2_contract_bundle.append(self.contract_table[message.cr].contract_bundle[message.consumer])

                min=0
                result_count = 150
                if round >= min and len(self.experiment_result.utility) <= result_count:
                    self.get_hub_result(min, round, message.cr, self.name, message.consumer.name,
                                        final_incentive, final_delay, result_count, Omega, Omega_prime, message.consumer)
            return M
        else:
            endTime = time.time()
            self.experiment_result.complete_payment += 1

            print("[{}] {} complete payGo : {}".format(round, self.name, self.experiment_result.complete_payment))

            min = 0
            result_count = 150

            temp = self.experiment_result.protocol_time[round]["lockTransfer"]
            self.experiment_result.protocol_time[round]["lockTransfer"] = message.locked_transfer_endtime - temp

            if round >= min and len(self.experiment_result.delay) <= result_count:
                self.get_result(min, round, message.cr, final_incentive, final_delay, endTime,
                                message.locked_transfer_endtime, result_count, Omega, Omega_prime)
            return M

    def get_result(self, min, round, cr, final_incentive, final_delay, endTime, locked_transfer_endtime, result_count, Omega, Omega_prime):
        # final_delay = (message.endTime - self.contract_table[message.cr].startTime) / time_meaningful_constant
        utility = contract_bundle(self.address).get_producer_utility(final_incentive, final_delay,Omega)

        if final_incentive == 0 :
            self.experiment_result.zero_incentive +=1

        self.experiment_result.complete_payment_round.append(round)
        self.experiment_result.delay.append(endTime - self.contract_table[cr].payment_startTime)
        self.experiment_result.contract_delay.append(final_delay)
        self.experiment_result.incentive.append(final_incentive)
        self.experiment_result.utility.append(utility)
        print("{} : accumulate round {}".format(self.name, len(self.experiment_result.delay)))
        if len(self.experiment_result.delay) == result_count:
            self.experiment_result.complete  = True
            if self.contract_table[cr].target.experiment_result.complete == True :
                processing_tx = self.experiment_result.complete_payment +self.contract_table[cr].target.experiment_result.complete_payment
                self.experiment_result.processing_tx = processing_tx
                self.contract_table[cr].target.experiment_result.processing_tx = processing_tx

                self.save_result(self)
                self.save_result(self.contract_table[cr].target)


    def save_result(self, node):
        wb = Workbook()
        # 파일 이름을 정하고, 데이터를 넣을 시트를 활성화합니다.
        sheet1 = wb.active
        file_name = node.name + '.xlsx'
        # 시트의 이름을 정합니다.
        sheet1.title = 'sampleSheet'
        # cell 함수를 이용해 넣을 데이터의 행렬 위치를 지정해줍니다.
        sheet1.cell(row=1, column=1).value = "delay"
        sheet1.cell(row=1, column=2).value = "Conditional payment"
        sheet1.cell(row=1, column=3).value = "incentive"
        sheet1.cell(row=1, column=4).value = "Utility"
        sheet1.cell(row=1, column=5).value = "queueing(h1)(count)"
        sheet1.cell(row=1, column=6).value = "queueing(h2)(count)"
        sheet1.cell(row=1, column=7).value = "capacity(h1)(after selection)"
        sheet1.cell(row=1, column=8).value = "capacity(h2)(after selection)"
        sheet1.cell(row=1, column=9).value = "balacnce(h1)"
        sheet1.cell(row=1, column=10).value = "balacnce(h2)"
        sheet1.cell(row=1, column=11).value = "selected_contract_delay(h1)"
        sheet1.cell(row=1, column=12).value = "selected_contract_delay(h2)"
        sheet1.cell(row=1, column=13).value = "protocol time(propose)"
        sheet1.cell(row=1, column=14).value = "protocol time(select)"
        sheet1.cell(row=1, column=15).value = "protocol time(lock)"
        sheet1.cell(row=1, column=16).value = "pending_payment_settle(p)"
        sheet1.cell(row=1, column=17).value = "pending_payment_settle(h1)"
        sheet1.cell(row=1, column=18).value = "pending_payment_settle(h2)"
        sheet1.cell(row=1, column=19).value = "zeroIncetive(count)"
        sheet1.cell(row=1, column=20).value = "onchainAccess(count)"
        sheet1.cell(row=1, column=21).value = "round"
        sheet1.cell(row=1, column=22).value = "tps"

        sheet1.cell(row=1, column=23).value = "onchain info(node) ->"
        sheet1.cell(row=1, column=24).value = "capacity"
        sheet1.cell(row=1, column=25).value = "balance"
        sheet1.cell(row=1, column=26).value = "selected_contract"
        sheet1.cell(row=1, column=27).value = "aux_contract"
        sheet1.cell(row=1, column=28).value = "pending_time"
        sheet1.cell(row=1, column=29).value = "pending_queue"

        for row_index in range(2, len(node.experiment_result.delay) + 2):
            index = row_index - 2
            complete_round = node.experiment_result.complete_payment_round[index]
            sheet1.cell(row=row_index, column=1).value = node.experiment_result.delay[index]
            sheet1.cell(row=row_index, column=2).value = node.experiment_result.contract_delay[index]
            sheet1.cell(row=row_index, column=3).value = node.experiment_result.incentive[index]
            sheet1.cell(row=row_index, column=4).value = node.experiment_result.utility[index]
            sheet1.cell(row=row_index, column=13).value = node.experiment_result.protocol_time[complete_round]["propose"]
            sheet1.cell(row=row_index, column=14).value = node.experiment_result.protocol_time[complete_round]["select"]
            sheet1.cell(row=row_index, column=15).value = node.experiment_result.protocol_time[complete_round]["lockTransfer"]
            sheet1.cell(row=row_index, column=21).value = complete_round

            # D는 1이 마지막 hub node, A는 2가 마지막 hub node
            if complete_round in node.experiment_result.pending_payment_settle:
                if 0 in node.experiment_result.pending_payment_settle[complete_round]:
                    sheet1.cell(row=row_index, column=16).value = \
                        node.experiment_result.pending_payment_settle[complete_round][0]["settle_time"]
                if 1 in node.experiment_result.pending_payment_settle[complete_round]:
                    sheet1.cell(row=row_index, column=17).value = \
                        node.experiment_result.pending_payment_settle[complete_round][1]["settle_time"]
                if 2 in node.experiment_result.pending_payment_settle[complete_round]:
                    sheet1.cell(row=row_index, column=18).value = \
                        node.experiment_result.pending_payment_settle[complete_round][2]["settle_time"]

            k = 5
            if complete_round in node.experiment_result.capacity :
                if node.name == "A" :
                    st = 1
                    dt =3
                    interval = 1
                else :
                    st =2
                    dt = 0
                    interval = -1

                for i in range(st, dt, interval) :
                    sheet1.cell(row=row_index, column=k).value = \
                        node.experiment_result.pending_queue[complete_round][i]
                    sheet1.cell(row=row_index, column=k+2).value = \
                        node.experiment_result.capacity[complete_round][i]
                    sheet1.cell(row=row_index, column=k+4).value = \
                        node.experiment_result.balance[complete_round][i]
                    sheet1.cell(row=row_index, column=k+6).value = \
                        node.experiment_result.selected_contract[complete_round][i]
                    k +=1

        sheet1.cell(row=2, column=19).value = node.experiment_result.zero_incentive
        sheet1.cell(row=2, column=20).value = node.experiment_result.onchain_access
        processing_time = time.time() - node.experiment_result.payGo_startTime
        print("processing_time : ", processing_time)
        print("processing_tx : ",  node.experiment_result.processing_tx)
        sheet1.cell(row=2, column=22).value = node.experiment_result.processing_tx / processing_time
        sheet1.cell(row=3, column=22).value = node.experiment_result.complete_payment / processing_time

        index = 2
        for i in node.experiment_result.onchain_access_node:
            sheet1.cell(row=index, column=23).value = node.experiment_result.onchain_access_node[i]
            sheet1.cell(row=index, column=24).value = node.experiment_result.onchain_access_capacity[i]
            sheet1.cell(row=index, column=25).value = node.experiment_result.onchain_access_balance[i]
            sheet1.cell(row=index, column=26).value = node.experiment_result.onchain_access_selected_contract[i][0]
            sheet1.cell(row=index, column=27).value = node.experiment_result.onchain_access_selected_contract[i][1]
            sheet1.cell(row=index, column=28).value = node.experiment_result.onchain_access_pendingTime[i]
            sheet1.cell(row=index, column=29).value = node.experiment_result.onchain_access_pendingQueue[i]
            index +=1

        print("*experiment_result : *", node.experiment_result.delay)
        result_delay = 0
        result_utlity = 0
        for i in range(len(node.experiment_result.delay)):
            result_delay += node.experiment_result.delay[i]
            result_utlity += node.experiment_result.utility[i]


        print("*{} delay* {}".format(node.name, result_delay / len(node.experiment_result.delay)))
        print("*{} utility {}*".format(node.name, result_utlity / len(node.experiment_result.delay)))
        print("test, result_delay : {}, len : {}".format(result_delay, len(node.experiment_result.delay)))
        print("test, result_utlity : {}, len : {}".format(result_utlity, len(node.experiment_result.delay)))
        print("*{} zero_incentive_count* {}".format(node.name, node.experiment_result.zero_incentive))

        wb.save(filename=file_name)

    def get_hub_result(self, min, round, cr, hub1, hub2, final_incentive, final_delay, result_count, Omega, Omega_prime, partner):
        utility = contract_bundle(self.address).get_producer_utility(final_incentive, final_delay,Omega)
        if final_incentive == 0 :
            self.experiment_result.zero_incentive +=1

        self.experiment_result.complete_payment_round.append(round)
        self.experiment_result.hub1.append(hub1)
        self.experiment_result.hub2.append(hub2)
        self.experiment_result.contract_delay.append(final_delay)
        self.experiment_result.incentive.append(final_incentive)
        self.experiment_result.utility.append(utility)

        if len(self.experiment_result.utility) == result_count:
            self.experiment_result.complete = True
            if partner.experiment_result.complete == True:
                self.save_hub_result(self)
                self.save_hub_result(partner)

    def save_hub_result(self, node):
        wb = Workbook()
        # 파일 이름을 정하고, 데이터를 넣을 시트를 활성화합니다.
        sheet1 = wb.active
        file_name = node.name + '(hub).xlsx'
        # 시트의 이름을 정합니다.
        sheet1.title = 'sampleSheet'
        # cell 함수를 이용해 넣을 데이터의 행렬 위치를 지정해줍니다.
        sheet1.cell(row=1, column=1).value = "hub(1)"
        sheet1.cell(row=1, column=2).value = "hub(2)"
        sheet1.cell(row=1, column=3).value = "contract delay"
        sheet1.cell(row=1, column=5).value = "incentive"
        sheet1.cell(row=1, column=6).value = "zero_incentive"
        sheet1.cell(row=1, column=7).value = "utility"
        sheet1.cell(row=1, column=8).value = "h1_contract_bundle"
        sheet1.cell(row=1, column=9).value = "h2_contract_bundle"

        for row_index in range(2, len(node.experiment_result.contract_delay) + 2):
            index = row_index - 2
            sheet1.cell(row=row_index, column=1).value = node.experiment_result.hub1[index]
            sheet1.cell(row=row_index, column=2).value = node.experiment_result.hub2[index]
            sheet1.cell(row=row_index, column=3).value = node.experiment_result.contract_delay[index]
            sheet1.cell(row=row_index, column=5).value = node.experiment_result.incentive[index]
            sheet1.cell(row=row_index, column=7).value = node.experiment_result.incentive[index]

            contract1 = ""
            contract2 = ""
            for i in range(len(node.experiment_result.h1_contract_bundle[index]["Incentive"])) :
                contract1 += "(" +str(node.experiment_result.h1_contract_bundle[index]["Incentive"][i][0] / contract_meaningful_incentive_constant) + ", " + \
                             str(node.experiment_result.h1_contract_bundle[index]["Delay"][i] / contract_meaningful_delay_constant) + "), "

            for i in range(len(node.experiment_result.h2_contract_bundle[index]["Incentive"])):
                contract2 += "(" + str(node.experiment_result.h2_contract_bundle[index]["Incentive"][i][0] / contract_meaningful_incentive_constant) + ", " + \
                             str(node.experiment_result.h2_contract_bundle[index]["Delay"][i] / contract_meaningful_delay_constant) + "), "

            sheet1.cell(row=row_index, column=8).value = contract1
            sheet1.cell(row=row_index, column=9).value = contract2

        sheet1.cell(row=2, column=6).value = node.experiment_result.zero_incentive

        wb.save(filename=file_name)


    def receive_unlockBP(self, message, round):
        state = self.partner[message.producer].unlock_BP(message.BP)
        self.partner[message.producer].pop_pending_payment(message.cr)
        self.partner[message.producer].update_average_capacity(message.finalAmount)
        consumer = self.contract_table[message.cr].consumer

        if self.contract_table[message.cr].target != self :
            final_incentive = message.finalAmount - self.contract_table[message.cr].amount
            delay = self.contract_table[message.cr].lock_time[1] - self.contract_table[message.cr].lock_time[0]
            self.update_contract_selection_standard(message.cr, delay, consumer, final_incentive)

    def update_contract_selection_standard(self,cr, delay, partner, final_incentive):
        queueing_record = self.partner[partner].get_queueing_record()
        # print("test ", self.contract_table[cr].lock_time)
        amount_ratio = self.contract_table[cr].amount_ratio

        if len(queueing_record) == 0:
            self.partner[partner].update_queueing_record(amount_ratio,
                                                         round(self.contract_table[cr].aready_reservation_delay, 3))
        # elif final_incentive == 0 :

            # if amount_ratio / delay <  queueing_record[0] / queueing_record[1] :
            #     self.partner[partner].update_queueing_record(amount_ratio, delay)
            #     print("update_contract_selection_standard", self.partner[partner].get_queueing_record())

    def register_secret_to_onchain(self, w3, cr, account, contract):
        secret = self.contract_table[cr].secret
        target = self.contract_table[cr].target
        endTime = self.contract_table[cr].endTime
        signature = self.contract_table[cr].signed_endTime
        packed_endTime = pack_data(['uint256'], [endTime])
        hash_endTime = '0x' + sha3(packed_endTime).hex()

        # print("target : ", target)
        # print("endTime : ", endTime)
        # print("signature : ", signature)
        # print("hash_endTime : ", hash_endTime)

        # w3, account, contract, *args = {secret, target, endTime, signature}
        result = register_secret(w3, account, contract, secret, target, endTime, signature)

        print("register secret : endTime -> {}".format(result[2]))

    def close_channel_to_onchain(self, w3, account, contract, partner):
        channel_identifier = self.partner[partner].channel_identifier
        BP = self.partner[partner].get_partner_BP()

        # print("channel_identifier : ", channel_identifier)
        # print("balance_hash : ", BP.balance_hash)
        # print("nonce : ", BP.message_data.nonce)
        # print("additional_hash : ", BP.additional_hash)
        # print("signature : ", BP.signature)

        # w3, account, contract, *args = {secret, target, endTime, signature}
        result = close_channel(w3, account, contract, account.address, channel_identifier, partner,
                               BP.balance_hash, BP.message_data.nonce, BP.additional_hash, BP.signature)

        print("close channel : channel_identifier,  closing_participant-> {}, {}".format(result[0], result[1]))
        print()

    def update_NonClosingBalanceProof_to_onchain(self, w3, account, contract, partner):

        channel_identifier = self.partner[partner].channel_identifier
        BP = self.partner[partner].get_partner_BP()
        packed_balance_proof = pack_data(['uint256', 'bytes32', 'uint256', 'bytes32', 'bytes'],
                                         [self.channel_identifier, BP.balance_hash, BP.message_data.nonce,
                                          BP.additional_hash, BP.signature])

        hashBP = '0x' + sha3(packed_balance_proof).hex()
        non_closing_signature = w3.eth.account.signHash(message_hash=hashBP, private_key=self.private_key)[
            'signature'].hex()

        # print("channel_identifier : ", channel_identifier)
        # print("balance_hash : ", BP.balance_hash)
        # print("nonce : ", BP.message_data.nonce)
        # print("additional_hash : ", BP.additional_hash)
        # print("signature : ", BP.signature)
        # print("non_closing_signature : ", non_closing_signature)

        result = update_NonClosingBalanceProof(w3, account, contract, channel_identifier, partner, account.address,
                                               BP.balance_hash, BP.message_data.nonce, BP.additional_hash, BP.signature,
                                               non_closing_signature)

        print("update_NonClosingBalanceProof : channel_identifier,  closing_participant, nonce-> {}, {}, {}".format(
            result[0], result[1], result[2]))

    def settle_channel_to_onchain(self, w3, account, contract, partner):
        channel_identifier = self.partner[partner].channel_identifier

        # [0] = transferred_amount, [1] = locked_amount, [2] = locksroot
        participant1 = self.partner[partner].get_state()
        participant2 = self.partner[partner].get_partner_state()

        # print("channel_identifier : ", channel_identifier)
        # print("participant1_transferred_amount : ", participant1[0])
        # print("participant1_locked_amount : ", participant1[1])
        # print("participant1_locksroot : ", participant1[2])
        # print("participant2_transferred_amount : ", participant2[0])
        # print("participant2_locked_amount : ", participant2[1])
        # print("participant2_locksroot : ", participant2[2])

        result = settle_channel(w3, account, contract, channel_identifier, account.address, participant1[0],
                                participant1[1],
                                participant1[2], partner, participant2[0], participant2[1], participant2[2])

        print(
            "settle_channel : channel_identifier,  participant1_return_amount, participant2_return_amount-> {}, {}, {}".format(
                result[0], result[1], result[2]))
        print()

    def unlock_to_onchain(self, w3, account, contract, partner):
        channel_identifier = self.partner[partner].channel_identifier
        merkle_tree_leaves = self.partner[partner].get_partner_leaves()

        # print("channel_identifier : ", channel_identifier)
        # print("merkle_tree_leaves : ", merkle_tree_leaves)
        # print("len merkle_tree_leaves :", len(merkle_tree_leaves))

        result = unlock(w3, account, contract, channel_identifier, account.address, partner, merkle_tree_leaves)

        print(
            "unlock to onchain : unlocked_amount(seccess), unsettled_amount(unsettled), returned_tokens(fail)-> {}, {}, {}".format(
                result[4], result[5], result[8]))
        print()

