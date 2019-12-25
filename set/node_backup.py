from message import fail, initContractPropose, contractPropose, contractSelect, LockedTransfer, sendSecret, resSecret, \
    revealSecret, resRevealSecret, completePayGo, waitContractSelect, selectionComplete, contractConfirm, contractReject
from algorism import contract_bundle
from settingParameter import on_chain_access_cost
from structure import contractTable, payer_result
from raiden.utils import sha3
from settingParameter import alpha, contract_meaningful_incentive_constant, contract_meaningful_delay_constant, \
    time_meaningful_constant
from contract import register_secret, close_channel, update_NonClosingBalanceProof, settle_channel, unlock
from raiden.utils.signing import pack_data

from openpyxl import Workbook
import os, math, time
import threading


class Node:
    # def __init__(self, roles = list()):
    #     self.cr = 0
    #     self.channel_state = roles
    #     self.address = roles[0].addrs[roles[0].i]
    #     self.partner = dict([(channelState.addrs[1-channelState.i], channelState) for channelState in roles])
    #     self.private_key = roles[0].sk
    #     self.contract_table = {}

    def __init__(self, w3, account, name, line, result):
        self.w3 = w3
        self.name = name
        self.line = line
        self.account = account
        self.address = account.address
        self.private_key = account.privateKey
        self.cr = 0
        self.channel_state = []
        self.partner = {}
        self.contract_table = {}
        self.experiment_result = result
        self.probability = {}

    def get_channel_probability(self, recipient):
        channel_probability = {}
        count = 0
        lock = threading.Lock()
        lock.acquire()
        try:
            probability = self.probability
        finally:
            lock.release()
        if 0 == len(probability.keys()):
            for partner in (self.partner.keys()):
                if partner != recipient:
                    channel_probability[count] = 1
                    count += 1
        else:
            channel_probability = probability

        # print("channel_probability : ", channel_probability)

        return channel_probability

    def create_channel_state(self, partner, channel_state):
        self.channel_state.append(channel_state)
        # self.partner[channel_state.addrs[1-channel_state.i]] = channel_state
        self.partner[partner] = channel_state

    def init_contract_propose(self, round, initiator, recipient, target, amount, secret, channel_probability,
                              payment_state):
        cr = '0x' + sha3(secret).hex()

        M = []
        if self.partner[recipient].check_balance(amount):
            bundle = contract_bundle(initiator.address).execute(channel_probability)
            message = contractPropose(cr, initiator, self, recipient, target, amount, bundle)

            if not cr in self.contract_table.keys():
                new_time = time.time()
                self.contract_table[cr] = contractTable(message)
                self.contract_table[cr].secret = secret
                self.contract_table[cr].payment_startTime = new_time
                self.experiment_result.protocol_time[round] = {"propose": new_time, "select": 0, "lockTransfer": 0}
                self.experiment_result.select_time[round] = {}

            # print("{} -> {} init_contract_propose".format(initiator.name, recipient.name))\
            self.contract_table[cr].contract_bundle[recipient] = bundle

            M.append(message)
        else:
            # print("[initator wow!!]{} amount : {}, balance : {}, reserve : {}, cr : {}, lock_cr : {}".format(
            #     self.name, amount, self.partner[recipient].test_get_balance(), self.partner[recipient].get_reserve_amount(),
            #     cr, self.partner[recipient].reserve_payment.keys()))

            if cr in self.contract_table.keys():
                if payment_state != "contractPropose":
                    return M

            message = initContractPropose(initiator, recipient, target, amount, secret, channel_probability)
            M.append(message)
            # accumulate_RTT = self.partner[recipient].RTT.get_accumulate_time()
            # time.sleep(accumulate_RTT)
            # print("init_contract_propose test 대기 {} : {}".format(self.name, accumulate_RTT))
            # return self.init_contract_propose(initiator, recipient, target, amount)
            # print("{} -> {} 재전송 init_contract_propose".format(initiator.name, recipient.name))

        return M

    def send_contract_propose(self, cr, initiator, producer, partner, target, amount, channel_probability,
                              receive_contract_bundle):
        # print("{} -> {} balance : {}".format(self.name, partner.name, self.partner[partner].test_get_balance()))
        if self.partner[partner].check_balance(amount):
            self.contract_table[cr].check_propose.append(partner)
            if partner == target:
                bundle = {"Incentive": 1, "Delay": 1}
            else:
                bundle = contract_bundle(self.address).execute(channel_probability)
            message = contractPropose(cr, initiator, self, partner, target, amount, bundle)
            self.contract_table[message.cr].contract_bundle[partner] = bundle

            # print("{} -> {} send_contract_propose".format(self.name, partner.name))

        else:
            # if initiator.name == "A" :
            #     print("[wow!!]{} amount : {}, balance : {}, reserve : {}, cr : {}, lock_cr : {}".format(
            #         self.name, amount, self.partner[partner].test_get_balance(), self.partner[partner].get_reserve_amount(),
            #         cr, self.partner[partner].reserve_payment.keys()))

            message = contractPropose(cr, initiator, producer, self, target, amount, receive_contract_bundle)

            # print("{} -> {} 재송 send_contract_propose".format(producer, self.name))

        return message

    def receive_contract_propose(self, message, payment_state, count, round):
        M = []
        if not message.cr in self.contract_table.keys():
            self.contract_table[message.cr] = contractTable(message)
        self.contract_table[message.cr].receive_contract_bundle[message.producer] = message.contract_bundle

        if count > self.contract_table[message.cr].selection_direction:
            self.contract_table[message.cr].selection_direction += 1
            return M

        self.contract_table[message.cr].selection_direction = 1
        # print("[{}, {}]{} -> {} send_contract_propose".format(message.initiator.name, round, message.producer.name, self.name))
        # print("[{}, {}]{} ->{} receive bundle : {}".format(message.initiator.name, round, message.producer.name, self.name, message.contract_bundle))
        # print("[{}] {} test_get_balance : {}".format(round, self.name, self.partner[message.producer].test_get_balance()))
        # print("state : ", self.contract_table[message.cr].state)
        if payment_state == "contractPropose":
            # check target
            if self.address == message.target.address:
                for recipient in self.contract_table[message.cr].receive_contract_bundle.keys():
                    if not recipient in self.contract_table[message.cr].check_propose :
                        M.append(self.send_target_contract_select(message.cr, recipient, 0, 0, 0, 0))
            else:
                for partner in self.partner.keys():
                    if partner != message.producer and partner.line != message.producer.line and not partner in \
                                                                                                     self.contract_table[
                                                                                                         message.cr].check_propose:
                        channel_probability = self.get_channel_probability(message.producer)
                        M.append(self.send_contract_propose(message.cr, message.initiator, message.producer, partner,
                                                            message.target,
                                                            message.amount, channel_probability,
                                                            message.contract_bundle))

        return M

    def send_target_contract_select(self, cr, recipient, incentive, delay, additional_incentive, additional_delay):
        table = self.contract_table[cr]
        self.contract_table[cr].check_propose.append(recipient)

        selected_incentive = table.receive_contract_bundle[recipient]["Incentive"]
        selected_delay = table.receive_contract_bundle[recipient]["Delay"]

        additional_incentive += incentive
        additional_delay += delay
        new_time = time.time()
        message = contractSelect(cr, table.initiator, self, recipient, table.target,
                                 selected_incentive, selected_delay, 0, additional_incentive,
                                 additional_delay, new_time)
        # path 설정
        message.path.append(self.name)

        return message

    def send_contract_select(self, cr, producer, consumer, incentive, delay, additional_incentive, additional_delay,
                             selected_index, path, propose_endTime):
        # 14 select
        table = self.contract_table[cr]
        self.contract_table[cr].state = "contractSelect"
        self.contract_table[cr].send_selected_contract.append(producer)
        try:
            selected_incentive = int(round(table.receive_contract_bundle[producer]["Incentive"][selected_index][0]))
            selected_delay = int(round(table.receive_contract_bundle[producer]["Delay"][selected_index]))
        except:
            print("contract error, receive contract bundle : {}".format(table.receive_contract_bundle[producer]))
            print("contract error selected index : ", selected_index)
            length = len(table.receive_contract_bundle[producer]["Incentive"])
            selected_incentive = int(round(table.receive_contract_bundle[producer]["Incentive"][length - 1][0]))
            selected_delay = int(round(table.receive_contract_bundle[producer]["Delay"][length - 1]))

        if selected_incentive >= on_chain_access_cost:
            print("incentive 초과 : {}".format(selected_incentive))
            message = fail("onchainCostExcess")
        else:
            additional_incentive += incentive
            additional_delay += delay
            message = contractSelect(cr, table.initiator, self, producer, table.target, selected_incentive,
                                     selected_delay,
                                     selected_index, additional_incentive, additional_delay, propose_endTime)

            # path 설정
            message.path.append(self.name)

        return message

    def contract_select(self, cr, selected_consumer, amount, producer, contract_bundle):
        index = -1

        if self.partner[selected_consumer].check_balance(amount):
            index = len(contract_bundle[producer]["Delay"]) - 1
        else:
            new_time = time.time()
            delay = 0
            total_amount = amount
            current_time = int(new_time * time_meaningful_constant)
            success = False

            lock = threading.Lock()
            with lock:
                pending_payment_bundle = self.partner[selected_consumer].pending_payment
                for item in list(pending_payment_bundle):
                    # [0] = locked amount, [1] = delay, [2] = startTime\
                    try:
                        pending_payment = pending_payment_bundle[item]
                    except:
                        print(
                            "[error 2] item : {}, pending_payment_bundle : {}".format(item, pending_payment_bundle))
                        continue
                    # new_delay = (pending_payment[2] + pending_payment[1]) - current_time
                    new_delay = (pending_payment[1] / contract_meaningful_delay_constant) - (current_time - pending_payment[2]) / time_meaningful_constant
                    # print("new_delay :", new_delay)
                    if new_delay < 0:
                        continue

                    delay += new_delay
                    total_amount -= pending_payment[0]
                    if total_amount <= 0:
                        total_amount = 0
                        success = True
                    # print("AAA")

                    if self.partner[selected_consumer].check_balance(total_amount):
                        # print("BBB")
                        for i in range(len(contract_bundle[producer]["Delay"]) - 1, -1, -1):
                            # delay * 5 -> 수정 가능
                            if delay <= contract_bundle[producer]["Delay"][i] / contract_meaningful_delay_constant:
                                index = i
                                print("성공1 ! index : {}, delay : {}, receive deay : {}".format(
                                        index, delay,contract_bundle[producer]["Delay"]))
                                self.contract_table[cr].redemption_delay = delay
                                success = True
                                break

                    if success: break

            # print("{} Contract Slection : {}".format(self.name, index))
        return index

    def receive_contract_select(self, message, count, payment_state, round):
        M = []
        if self.contract_table[message.cr].state == "contractSelect":
            return M

        if payment_state == "lockedTransfer":
            return M

        if not message.consumer in self.contract_table[message.cr].temp_consumer:
            self.contract_table[message.cr].temp_selected_contract[message.consumer] = message.incentive, message.delay
            self.contract_table[message.cr].temp_additional_contract[
                message.consumer] = message.additional_incentive, message.additional_delay
            self.contract_table[message.cr].temp_selected_index[message.consumer] = message.selected_index
            self.contract_table[message.cr].temp_consumer.append(message.consumer)
        # else :
        #     print("[{}, {}]{}' with {} balance : {}, reserve amount : {}".format(
        #         message.initiator.name, round, self.name, message.consumer.name, self.partner[message.consumer].test_get_balance(),
        #         self.partner[message.consumer].get_reserve_amount()
        #     ))

        if count > self.contract_table[message.cr].selection_direction:
            self.contract_table[message.cr].selection_direction += 1
            return M
        # print("[{}, {}]{} <- {} send_contract_select".format(message.initiator.name, round, self.name,
        #                                                      message.consumer.name))

        # 나중을 위해 초기화
        self.contract_table[message.cr].selection_direction = 1

        table = self.contract_table[message.cr]
        select = 10000000000  # infinite value
        selected_consumer = ""

        if self == message.initiator:
            temp_consumer = []
            for i in range(len(table.temp_consumer)):
                selected_incentive = table.temp_selected_contract[table.temp_consumer[i]][0]
                additional_incentive = table.temp_additional_contract[table.temp_consumer[i]][0]
                if self.partner[table.temp_consumer[i]].check_balance(
                        table.amount + selected_incentive + additional_incentive):
                    temp_consumer.append(table.temp_consumer[i])
            table.temp_consumer = temp_consumer

            if len(temp_consumer) == 0:
                M.append(message)
                # print("[{}] {} initiator channel bankrupt".format(round, self.name))
                return M

        for i in range(len(table.temp_consumer)):
            selected_delay = table.temp_selected_contract[table.temp_consumer[i]][1]
            additional_delay = table.temp_additional_contract[table.temp_consumer[i]][1]
            if selected_delay + additional_delay < select:
                select = selected_delay + additional_delay
                selected_consumer = table.temp_consumer[i]

        selected_contract = table.temp_selected_contract[selected_consumer]
        additional_contract = table.temp_additional_contract[selected_consumer]
        selected_index = table.temp_selected_index[selected_consumer]

        self.contract_table[message.cr].selected_contract = selected_contract
        self.contract_table[message.cr].additional_contract = additional_contract
        self.contract_table[message.cr].selected_index = selected_index
        self.contract_table[message.cr].consumer = selected_consumer
        # if self.contract_table[message.cr].consumer != "" and message.target != message.consumer:
        #     if self.contract_table[message.cr].consumer != selected_consumer :
        #         M.append(self.send_contract_reject(message.cr, self.contract_table[message.cr].consumer))
        #     else :
        #         M.append(self.send_contract_reject(message.cr, message.consumer))

        if message.consumer != message.target:
            probability = {}
            for i in range(len(table.contract_bundle[message.consumer]["Incentive"])):
                probability[i] = 0
            for i in table.temp_selected_index.values():
                probability[i] += 1

            lock = threading.Lock()
            lock.acquire()
            try:
                self.probability = probability
            finally:
                lock.release()

            # print("수정 probability : ", self.probability)

        amount = self.contract_table[message.cr].amount + selected_contract[0] + additional_contract[0]

        if self == message.initiator:
            a = self.reserve_payment(message.cr, selected_consumer, amount)
            new_time = time.time()
            self.contract_table[message.cr].state = "contractSelect"
            temp = self.experiment_result.protocol_time[round]["propose"]
            self.experiment_result.protocol_time[round]["propose"] = message.propose_endTime - temp
            self.experiment_result.protocol_time[round]["select"] =  new_time- message.propose_endTime
            M.append(self.send_secret(message.cr, message.initiator, message.target))
            M.append(self.send_contract_confirm(message.cr, selected_consumer))
        else:
            sucess = 0
            for recipient in self.contract_table[message.cr].receive_contract_bundle.keys():
                if sucess == 0 :
                    selected_index = self.contract_select(message.cr, selected_consumer, amount,
                                                          recipient, table.receive_contract_bundle)
                else :
                    selected_index = len(table.receive_contract_bundle[recipient]["Delay"]) -1

                if selected_index == -1:
                    # if not self.line in message.initiator.experiment_result.select_time[round] :
                    #     message.initiator.experiment_result.select_time[round][self.line] = time.time()
                    continue
                else:
                    # if self.name in message.initiator.experiment_result.select_time[round]:
                    #     temp = message.initiator.experiment_result.select_time[round][self.name]
                    #     message.initiator.experiment_result.select_time[round][self.name] = time.time() - temp
                    sucess +=1
                    M.append(self.send_contract_select(message.cr, recipient, message.consumer, selected_contract[0],
                                                       selected_contract[1], additional_contract[0],
                                                       additional_contract[1], selected_index, message.path,
                                                       message.propose_endTime))
            if sucess == 0:
                M.append(message)

        return M

    def reserve_payment(self, cr, consumer, amount):
        lock = threading.Lock()
        with lock:
            reserve_payment = self.partner[consumer].reserve_payment

        if not cr in reserve_payment:
            self.partner[consumer].reserve_payment[cr] = amount


    def send_contract_confirm(self, cr, selected_consumer):
        message = contractConfirm(cr, self, selected_consumer)

        return message

    def receive_contract_confirm(self, message, round):
        # print("[{}, {}] {}' receive_contract_confirm ".format(round, self.contract_table[message.cr].initiator.name, self.name))
        M = []
        consumer = self.contract_table[message.cr].consumer
        incentive = self.contract_table[message.cr].selected_contract[0] + \
                    self.contract_table[message.cr].additional_contract[0]
        amount = self.contract_table[message.cr].amount + incentive
        a = self.reserve_payment(message.cr, consumer, amount)

        if consumer != self.contract_table[message.cr].target :
            M.append(self.send_contract_confirm(message.cr, consumer))
        return M


    def send_secret(self, cr, initiator, target):
        secret = self.contract_table[cr].secret
        # send Secret
        message = sendSecret(cr, secret, initiator, target)
        return message

    # message.cr, initiator, 0, recipient, target, amount, sha3(message.secret), 0
    def send_locked_transfer(self, cr, initiator, producer, recipient, target, amount, secretHash, receive_BP, r):
        table = self.contract_table[cr]
        result_incentive = (table.selected_contract[0] + table.additional_contract[0])
        result_delay = table.selected_contract[1] + table.additional_contract[1]
        selected_index = table.selected_index

        if selected_index <= 0:
            aux_incentive = result_incentive
            aux_delay = result_delay
        else:
            if round(table.contract_bundle[recipient]["Incentive"][selected_index - 1][0]) == 0:
                aux_incentive = result_incentive
                aux_delay = result_delay
            else:
                aux_incentive = int(round(table.contract_bundle[recipient]["Incentive"][selected_index - 1][0]) +
                                    table.additional_contract[0])
                aux_delay = int(
                    round(table.contract_bundle[recipient]["Delay"][selected_index - 1]) + table.additional_contract[1])

        lock = threading.Lock()
        with lock:
            reserve_payment = self.partner[recipient].reserve_payment

        if not cr in reserve_payment:
            reserve_amount = 0
        else:
            reserve_amount = reserve_payment[cr]
        new_time = time.time()
        if self.partner[recipient].check_balance(self.contract_table[cr].amount + result_incentive, reserve_amount):
            startTime = int(new_time * time_meaningful_constant)
            self.contract_table[cr].startTime = startTime

            # balance data 갱신
            expiration = startTime + result_delay + alpha
            lock = threading.Lock()
            with lock:
                try:
                    self.partner[recipient].reserve_payment.pop(cr)
                except:
                    print("[hummm error]{} amount : {}, balance : {}, reserve : {}, cr : {}, lock_cr : {}".format(
                        self.name, self.contract_table[cr].amount + result_incentive,
                        self.partner[recipient].test_get_balance(), self.partner[recipient].get_reserve_amount(),
                        cr, self.partner[recipient].reserve_payment.keys()))

                BP = self.partner[recipient].create_BP(self.w3, cr, initiator.address, target.address, secretHash,
                                                       amount, expiration,
                                                       (result_incentive, result_delay), (aux_incentive, aux_delay),
                                                       startTime)

            message = LockedTransfer(cr, self, recipient, BP, target, initiator)

            if r in initiator.experiment_result.pending_payment_settle:
                if self.line in initiator.experiment_result.pending_payment_settle[r] :
                    temp = initiator.experiment_result.pending_payment_settle[r][self.line]["settle_time"]
                    initiator.experiment_result.pending_payment_settle[r][self.line]["settle_time"] = new_time - temp
        else:
            if not r in initiator.experiment_result.pending_payment_settle :
                initiator.experiment_result.pending_payment_settle[r] = {}
                initiator.experiment_result.pending_payment_settle[r][self.line] = \
                    {"count": len(self.partner[recipient].pending_payment.keys()), "settle_time": new_time}
            elif not self.line in initiator.experiment_result.pending_payment_settle[r] :
                initiator.experiment_result.pending_payment_settle[r][self.line] = \
                    {"count": len(self.partner[recipient].pending_payment.keys()), "settle_time": new_time}

            if self == initiator:
                message = resSecret(cr, initiator, table.secret)
            else:
                message = LockedTransfer(cr, producer, self, receive_BP, target, initiator)
        return message

    def receive_locked_transfer(self, message, round):
        if not message.cr in self.partner[message.producer].pending_payment.keys():
            pending_payment = self.partner[message.producer].locked_BP(message.BP)
            self.partner[message.producer].pending_payment[message.cr] = pending_payment
            self.contract_table[message.cr].producer = message.producer

        # print("[{}, {}]{} -> {} locked transfer".format(message.initiator.name, round, message.producer.name, self.name))
        # print("{} -> {} start Time :  {}".format(message.producer.name, self.name, message.BP.message_data.start_time))

        if self.address == message.target.address:
            M = [self.target_reveal_secret(message.cr, message.producer, message.consumer)]
            return M
        else:
            partner = self.contract_table[message.cr].consumer
            amount = self.contract_table[message.cr].amount
            secrethash = message.BP.message_data.secrethash
            M = [self.send_locked_transfer(message.cr, message.initiator, message.producer, partner, message.target,
                                           amount, secrethash, message.BP, round)]
            return M

    def receive_secret(self, message):
        self.contract_table[message.cr].secret = message.secret

        # print("{} -> {} send secret".format(message.initiator.name, self.name))

        M = [self.send_res_secret(message.cr, message.initiator, message.secret)]
        return M

    def send_res_secret(self, cr, initiator, secret):
        message = resSecret(cr, initiator, secret)
        return message

    def recevie_res_secret(self, message, round):
        initiator = message.initiator
        recipient = self.contract_table[message.cr].consumer
        target = self.contract_table[message.cr].target
        amount = self.contract_table[message.cr].amount
        # print("round : ", round)
        new_time = time.time()
        self.experiment_result.protocol_time[round]["lockTransfer"] = new_time
        M = [self.send_locked_transfer(message.cr, initiator, 0, recipient, target, amount, sha3(message.secret), 0, round)]
        return M

    def target_reveal_secret(self, cr, producer, consumer):
        if self.contract_table[cr].secret != 0:
            new_time = time.time()
            endTime = int(new_time * time_meaningful_constant)
            # print("test endtime :", endTime)
            packed_endTime = pack_data(['uint256'], [endTime])
            hash_endTime = '0x' + sha3(packed_endTime).hex()
            signed_endTime = self.w3.eth.account.signHash(message_hash=hash_endTime, private_key=self.private_key)

            message = revealSecret(cr, self.contract_table[cr].secret, endTime, signed_endTime['signature'].hex(),
                                   producer, consumer, new_time)

            self.contract_table[cr].endTime = endTime
            self.contract_table[cr].signed_endTime = signed_endTime['signature'].hex()

            return message

    def reveal_secret(self, cr, producer, consumer):
        if self.contract_table[cr].secret != 0 or self.contract_table[cr].signed_endTime != 0:
            endTime = self.contract_table[cr].endTime
            signed_endTime = self.contract_table[cr].signed_endTime
            message = revealSecret(cr, self.contract_table[cr].secret, endTime, signed_endTime, producer, consumer)

            print("endTime : ", endTime)
            print("signed_endTime : ", signed_endTime)

            return message

    def receive_reveal_secret(self, message, round, total_round):
        self.contract_table[message.cr].secret = message.secret
        self.contract_table[message.cr].endTime = message.endTime
        self.contract_table[message.cr].signed_endTime = message.signed_endTime
        initiator = self.contract_table[message.cr].initiator
        BP, finalAmount, final_incentive, final_delay = self.partner[message.consumer].unlock(self.w3, message.cr,
                                                                                              message.secret,
                                                                                              message.endTime)
        final_incentive = final_incentive / contract_meaningful_incentive_constant
        M = [resRevealSecret(message.cr, message.producer, message.consumer, BP, finalAmount)]
        producer = self.contract_table[message.cr].producer

        # print("{} -> {} reveal_secret".format(message.consumer.name, self.name))
        # print("[{}, {}]{} end time : {}, start time {}".format(initiator.name, round, self.name,
        #                                                       message.endTime, self.contract_table[message.cr].startTime))

        if initiator != self:
            M.append(revealSecret(message.cr, self.contract_table[message.cr].secret, message.endTime,
                                  message.signed_endTime, producer, message.producer, message.locked_transfer_endtime))

            if self.line - initiator.line == 1:
                utility = contract_bundle(self.address).get_producer_utility(final_incentive, final_delay,
                                                                             contract_meaningful_incentive_constant,
                                                                             contract_meaningful_delay_constant)
                self.experiment_result.total_utility.append(utility)
                min=0
                result_count = 200
                if round >= min:
                    self.get_hub_result(min, round, message.cr, self.name, message.consumer.name,
                                        final_incentive, final_delay, result_count)
            return M
        else:
            endTime = time.time()
            self.experiment_result.complete_payment += 1
            print("[{}] {} complete payGo : {}".format(round, self.name, self.experiment_result.complete_payment))
            utility = contract_bundle(self.address).get_producer_utility(final_incentive, final_delay,
                                                                         contract_meaningful_incentive_constant,
                                                                         contract_meaningful_delay_constant)
            self.experiment_result.total_utility.append(utility)

            min = 0
            result_count = 200

            temp = self.experiment_result.protocol_time[round]["lockTransfer"]
            self.experiment_result.protocol_time[round]["lockTransfer"] = message.locked_transfer_endtime - temp

            if round >= min and self.name == "A":
                self.get_result(min, round, message.cr, final_incentive, final_delay, endTime,
                                message.locked_transfer_endtime, result_count)
            return M

    def get_result(self, min, round, cr, final_incentive, final_delay, endTime, locked_transfer_endtime, result_count):
        # final_delay = (message.endTime - self.contract_table[message.cr].startTime) / time_meaningful_constant
        utility = contract_bundle(self.address).get_producer_utility(final_incentive, final_delay,
                                                                     contract_meaningful_incentive_constant,
                                                                     contract_meaningful_delay_constant)

        print("[{}]{} final incetive : {}".format(round, self.name, final_incentive))
        print("[{}]{} final delay : {}".format(round, self.name, final_delay))
        print("[{}]{} payGo complete time : {}".format(round, self.name,
                                                       endTime - self.contract_table[cr].payment_startTime))
        print("[{}]{} payGo complete utility : {}".format(round, self.name, utility))

        if final_incentive == 0:
            self.experiment_result.zero_incentive += 1
            self.experiment_result.zero_incentive_delay[round] = final_delay

        self.experiment_result.complete_payment_round.append(round)
        self.experiment_result.delay.append(endTime - self.contract_table[cr].payment_startTime)
        self.experiment_result.contract_delay.append(final_delay)
        self.experiment_result.utility.append(utility)
        print("{} : accumulate round {}".format(self.name, len(self.experiment_result.delay)))
        if len(self.experiment_result.delay) == result_count:
            wb = Workbook()
            # 파일 이름을 정하고, 데이터를 넣을 시트를 활성화합니다.
            sheet1 = wb.active
            file_name = 'result(payer).xlsx'
            # 시트의 이름을 정합니다.
            sheet1.title = 'sampleSheet'
            # cell 함수를 이용해 넣을 데이터의 행렬 위치를 지정해줍니다.
            sheet1.cell(row=1, column=1).value = "delay"
            sheet1.cell(row=1, column=2).value = "contract delay"
            sheet1.cell(row=1, column=3).value = "Utility"
            sheet1.cell(row=1, column=4).value = "protocol time(propose)"
            sheet1.cell(row=1, column=5).value = "protocol time(select)"
            sheet1.cell(row=1, column=6).value = "protocol time(lock)"
            sheet1.cell(row=1, column=7).value = "pending_payment_settle(A)(count)"
            sheet1.cell(row=1, column=8).value = "pending_payment_settle(A)(settle time)"
            sheet1.cell(row=1, column=9).value = "pending_payment_settle(B)(count)"
            sheet1.cell(row=1, column=10).value = "pending_payment_settle(B)(settle time)"
            sheet1.cell(row=1, column=11).value = "pending_payment_settle(C)(count)"
            sheet1.cell(row=1, column=12).value = "pending_payment_settle(C)(settle time)"
            sheet1.cell(row=1, column=13).value = "zeroIncetive(count)"
            sheet1.cell(row=1, column=14).value = "zeroIncetive(settle delay)"
            sheet1.cell(row=1, column=15).value = "total_Utility"
            sheet1.cell(row=1, column=16).value = "round"
            # sheet1.cell(row=1, column=17).value = "select(A)"
            # sheet1.cell(row=1, column=18).value = "select(B)"
            # sheet1.cell(row=1, column=19).value = "select(C)"

            for row_index in range(2, len(self.experiment_result.delay)+2):
                index = row_index - 2
                complete_round = self.experiment_result.complete_payment_round[index]
                sheet1.cell(row=row_index, column=16).value = complete_round
                sheet1.cell(row=row_index, column=1).value = self.experiment_result.delay[index]
                sheet1.cell(row=row_index, column=2).value = self.experiment_result.contract_delay[index]
                sheet1.cell(row=row_index, column=3).value = self.experiment_result.utility[index]
                sheet1.cell(row=row_index, column=4).value = self.experiment_result.protocol_time[complete_round]["propose"]
                sheet1.cell(row=row_index, column=5).value = self.experiment_result.protocol_time[complete_round]["select"]
                sheet1.cell(row=row_index, column=6).value = self.experiment_result.protocol_time[complete_round]["lockTransfer"]

                if complete_round in self.experiment_result.pending_payment_settle :
                    if 0 in self.experiment_result.pending_payment_settle[complete_round] :
                        sheet1.cell(row=row_index, column=7).value = \
                            self.experiment_result.pending_payment_settle[complete_round][0]["count"]
                        sheet1.cell(row=row_index, column=8).value = \
                            self.experiment_result.pending_payment_settle[complete_round][0]["settle_time"]
                    if 1 in self.experiment_result.pending_payment_settle[complete_round] :
                        sheet1.cell(row=row_index, column=9).value = \
                            self.experiment_result.pending_payment_settle[complete_round][1]["count"]
                        sheet1.cell(row=row_index, column=10).value = \
                            self.experiment_result.pending_payment_settle[complete_round][1]["settle_time"]
                    if 2 in self.experiment_result.pending_payment_settle[complete_round] :
                        sheet1.cell(row=row_index, column=11).value = \
                            self.experiment_result.pending_payment_settle[complete_round][2]["count"]
                        sheet1.cell(row=row_index, column=12).value = \
                            self.experiment_result.pending_payment_settle[complete_round][2]["settle_time"]
                    # if "A" in self.experiment_result.select_time[complete_round] :
                    #     sheet1.cell(row=row_index, column=11).value = \
                    #         self.experiment_result.pending_payment_settle[complete_round][2]["count"]
                    #     sheet1.cell(row=row_index, column=12).value = \
                    #         self.experiment_result.pending_payment_settle[complete_round][2]["settle_time"]


            sheet1.cell(row=2, column=13).value = self.experiment_result.zero_incentive
            for r in self.experiment_result.zero_incentive_delay.keys():
                if r in self.experiment_result.complete_payment_round :
                    index = self.experiment_result.complete_payment_round.index(r)
                    sheet1.cell(row=2 + index, column=14).value = self.experiment_result.zero_incentive_delay[r]

            total_U = 0
            for i in range(len(self.experiment_result.total_utility)):
                total_U += self.experiment_result.total_utility[i]
            sheet1.cell(row=2, column=15).value = total_U / len(self.experiment_result.total_utility)

            print("*experiment_result : *", self.experiment_result.delay)
            result_delay = 0
            result_utlity = 0
            for i in range(len(self.experiment_result.delay)):
                result_delay += self.experiment_result.delay[i]
                result_utlity += self.experiment_result.utility[i]

            print("*{} delay* {}".format(self.name, result_delay / len(self.experiment_result.delay)))
            print("*{} utility {}*".format(self.name, result_utlity / len(self.experiment_result.delay)))
            print("test, result_delay : {}, len : {}".format(result_delay, len(self.experiment_result.delay)))
            print("test, result_utlity : {}, len : {}".format(result_utlity, len(self.experiment_result.delay)))
            print("*{} zero_incentive_count* {}".format(self.name, self.experiment_result.zero_incentive))

            sheet1.cell(row=len(self.experiment_result.delay) + 3, column=1).value = result_delay / len(self.experiment_result.delay)
            sheet1.cell(row=len(self.experiment_result.delay) + 3, column=3).value = result_utlity / len(self.experiment_result.delay)
            wb.save(filename=file_name)

    def get_hub_result(self, min, round, cr, hub1, hub2, final_incentive, final_delay, result_count):
        # final_delay = (message.endTime - self.contract_table[message.cr].startTime) / time_meaningful_constant
        utility = contract_bundle(self.address).get_producer_utility(final_incentive, final_delay,
                                                                     contract_meaningful_incentive_constant,
                                                                     contract_meaningful_delay_constant)
        if final_incentive == 0:
            self.experiment_result.zero_incentive += 1
            self.experiment_result.zero_incentive_delay[round] = final_delay

        self.experiment_result.complete_payment_round.append(round)
        self.experiment_result.hub1.append(hub1)
        self.experiment_result.hub2.append(hub2)
        self.experiment_result.contract_delay.append(final_delay)
        self.experiment_result.utility.append(utility)
        print("{} : accumulate round {}".format(self.name, len(self.experiment_result.contract_delay)))

        if len(self.experiment_result.contract_delay) == result_count:
            wb = Workbook()
            # 파일 이름을 정하고, 데이터를 넣을 시트를 활성화합니다.
            sheet1 = wb.active
            file_name = 'result(hub).xlsx'
            # 시트의 이름을 정합니다.
            sheet1.title = 'sampleSheet'
            # cell 함수를 이용해 넣을 데이터의 행렬 위치를 지정해줍니다.
            sheet1.cell(row=1, column=1).value = "hub(1)"
            sheet1.cell(row=1, column=2).value = "hub(2)"
            sheet1.cell(row=1, column=3).value = "contract delay"
            sheet1.cell(row=1, column=4).value = "Utility"
            sheet1.cell(row=1, column=5).value = "zeroIncetive(count)"
            sheet1.cell(row=1, column=6).value = "zeroIncetive(settle delay)"
            sheet1.cell(row=1, column=7).value = "total_Utility"

            for row_index in range(2, len(self.experiment_result.contract_delay)+2):
                index = row_index - 2
                sheet1.cell(row=row_index, column=1).value = self.experiment_result.hub1[index]
                sheet1.cell(row=row_index, column=2).value = self.experiment_result.hub2[index]
                sheet1.cell(row=row_index, column=3).value = self.experiment_result.contract_delay[index]
                sheet1.cell(row=row_index, column=4).value = self.experiment_result.utility[index]

            sheet1.cell(row=2, column=5).value = self.experiment_result.zero_incentive
            for r in self.experiment_result.zero_incentive_delay.keys():
                if r in self.experiment_result.complete_payment_round:
                    index = self.experiment_result.complete_payment_round.index(r)
                    sheet1.cell(row=2 + index, column=6).value = self.experiment_result.zero_incentive_delay[r]

            total_U = 0
            for i in range(len(self.experiment_result.total_utility)):
                total_U += self.experiment_result.total_utility[i]
            sheet1.cell(row=2, column=7).value = total_U / len(self.experiment_result.total_utility)

            wb.save(filename=file_name)

    def receive_unlockBP(self, message, round):
        state = self.partner[message.producer].unlock_BP(message.BP)
        lock = threading.Lock()
        lock.acquire()
        try:
            self.partner[message.producer].pending_payment.pop(message.cr)
        finally:
            lock.release()

        table = self.contract_table[message.cr]
        # if self != table.target :
        #     print("[{}]{} balance : [{}, {}],[{}, {}]".format(round, self.name, table.producer.name,
        #                                           self.partner[table.producer].test_get_balance(), table.consumer.name,
        #                                           self.partner[table.consumer].test_get_balance()))
        # else :
        #     print("[{}]{} balance : [{}, {}]".format(round, self.name, table.producer.name,
        #                                           self.partner[table.producer].test_get_balance()))
        # print("{} -> {} receive unlockBP".format(message.producer.name, self.name))
        # print("[{}]{}' with {} now balance : {}, reservation : {}".format(
        #     round, self.name, message.producer.name, self.partner[message.producer].test_get_balance(),
        #     self.partner[message.producer].get_reserve_amount()))

        M = [completePayGo(self.name)]
        return M

        # if message.producer == self.contract_table[message.cr].initiator :
        #     M = [completePayGo(message.newRTT)]
        #     return M

    def register_secret_to_onchain(self, w3, cr, account, contract):
        secret = self.contract_table[cr].secret
        target = self.contract_table[cr].target
        endTime = self.contract_table[cr].endTime
        signature = self.contract_table[cr].signed_endTime
        packed_endTime = pack_data(['uint256'], [endTime])
        hash_endTime = '0x' + sha3(packed_endTime).hex()

        # print("target : ", target)
        # print("endTime : ", endTime)
        # print("signature : ", signature)
        # print("hash_endTime : ", hash_endTime)

        # w3, account, contract, *args = {secret, target, endTime, signature}
        result = register_secret(w3, account, contract, secret, target, endTime, signature)

        print("register secret : endTime -> {}".format(result[2]))

    def close_channel_to_onchain(self, w3, account, contract, partner):
        channel_identifier = self.partner[partner].channel_identifier
        BP = self.partner[partner].get_partner_BP()

        # print("channel_identifier : ", channel_identifier)
        # print("balance_hash : ", BP.balance_hash)
        # print("nonce : ", BP.message_data.nonce)
        # print("additional_hash : ", BP.additional_hash)
        # print("signature : ", BP.signature)

        # w3, account, contract, *args = {secret, target, endTime, signature}
        result = close_channel(w3, account, contract, account.address, channel_identifier, partner,
                               BP.balance_hash, BP.message_data.nonce, BP.additional_hash, BP.signature)

        print("close channel : channel_identifier,  closing_participant-> {}, {}".format(result[0], result[1]))
        print()

    def update_NonClosingBalanceProof_to_onchain(self, w3, account, contract, partner):

        channel_identifier = self.partner[partner].channel_identifier
        BP = self.partner[partner].get_partner_BP()
        packed_balance_proof = pack_data(['uint256', 'bytes32', 'uint256', 'bytes32', 'bytes'],
                                         [self.channel_identifier, BP.balance_hash, BP.message_data.nonce,
                                          BP.additional_hash, BP.signature])

        hashBP = '0x' + sha3(packed_balance_proof).hex()
        non_closing_signature = w3.eth.account.signHash(message_hash=hashBP, private_key=self.private_key)[
            'signature'].hex()

        # print("channel_identifier : ", channel_identifier)
        # print("balance_hash : ", BP.balance_hash)
        # print("nonce : ", BP.message_data.nonce)
        # print("additional_hash : ", BP.additional_hash)
        # print("signature : ", BP.signature)
        # print("non_closing_signature : ", non_closing_signature)

        result = update_NonClosingBalanceProof(w3, account, contract, channel_identifier, partner, account.address,
                                               BP.balance_hash, BP.message_data.nonce, BP.additional_hash, BP.signature,
                                               non_closing_signature)

        print("update_NonClosingBalanceProof : channel_identifier,  closing_participant, nonce-> {}, {}, {}".format(
            result[0], result[1], result[2]))

    def settle_channel_to_onchain(self, w3, account, contract, partner):
        channel_identifier = self.partner[partner].channel_identifier

        # [0] = transferred_amount, [1] = locked_amount, [2] = locksroot
        participant1 = self.partner[partner].get_state()
        participant2 = self.partner[partner].get_partner_state()

        # print("channel_identifier : ", channel_identifier)
        # print("participant1_transferred_amount : ", participant1[0])
        # print("participant1_locked_amount : ", participant1[1])
        # print("participant1_locksroot : ", participant1[2])
        # print("participant2_transferred_amount : ", participant2[0])
        # print("participant2_locked_amount : ", participant2[1])
        # print("participant2_locksroot : ", participant2[2])

        result = settle_channel(w3, account, contract, channel_identifier, account.address, participant1[0],
                                participant1[1],
                                participant1[2], partner, participant2[0], participant2[1], participant2[2])

        print(
            "settle_channel : channel_identifier,  participant1_return_amount, participant2_return_amount-> {}, {}, {}".format(
                result[0], result[1], result[2]))
        print()

    def unlock_to_onchain(self, w3, account, contract, partner):
        channel_identifier = self.partner[partner].channel_identifier
        merkle_tree_leaves = self.partner[partner].get_partner_leaves()

        # print("channel_identifier : ", channel_identifier)
        # print("merkle_tree_leaves : ", merkle_tree_leaves)
        # print("len merkle_tree_leaves :", len(merkle_tree_leaves))

        result = unlock(w3, account, contract, channel_identifier, account.address, partner, merkle_tree_leaves)

        print(
            "unlock to onchain : unlocked_amount(seccess), unsettled_amount(unsettled), returned_tokens(fail)-> {}, {}, {}".format(
                result[4], result[5], result[8]))
        print()

